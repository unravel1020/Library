from sqlalchemy import create_engine, and_, func, or_
from sqlalchemy.orm import sessionmaker
from models import (
    Base, SeatType, ReservationStatus, Reminder, Reservation, Library,
    ReservationStats, Area, Seat, SystemLog, LogLevel, UserStatus,
    Blacklist, PointsHistory, Notification, NotificationTemplate, NotificationType,
    NotificationStatus, Report, ReportType, ReportFormat, SystemConfig,
    LibrarySchedule, ReservationRule, User
)
from datetime import datetime, timedelta
import hashlib
import json
import csv
import os
import shutil
from typing import List, Dict, Any, Optional
import secrets
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import requests
from dotenv import load_dotenv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO

engine = create_engine('sqlite:///library.db', connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)

def init_db():
    Base.metadata.create_all(bind=engine)

def get_seat_status(library_id):
    session = SessionLocal()
    try:
        # Use Library model instead of SeatStatus
        library = session.query(Library).filter_by(id=library_id).first()
        return library
    finally:
        session.close()

def update_seat_status(library_id, available, occupied, reserved):
    session = SessionLocal()
    try:
        # Use Library model instead of SeatStatus
        library = session.query(Library).filter_by(id=library_id).first()
        if library:
            library.available_seats = available
            library.occupied_seats = occupied
            library.reserved_seats = reserved
            session.commit()
        return library # Return the updated library object
    finally:
        session.close()

def get_all_libraries():
    session = SessionLocal()
    libraries = session.query(Library).all()
    session.close()
    return libraries

def add_library(id, name, location):
    session = SessionLocal()
    library = Library(
        id=id,
        name=name,
        location=location,
        total_seats=0,
        available_seats=0,
        occupied_seats=0,
        reserved_seats=0
    )
    session.add(library)
    session.commit()
    session.close()
    return library

def check_reservation_conflict(seat_id: int, start_time: datetime, end_time: datetime):
    session = SessionLocal()
    conflict = session.query(Reservation).filter(
        and_(
            Reservation.seat_id == seat_id,
            Reservation.status == ReservationStatus.ACTIVE,
            Reservation.start_time < end_time,
            Reservation.end_time > start_time
        )
    ).first()
    session.close()
    return conflict is not None

def check_reservation_duration(seat_id: int, start_time: datetime, end_time: datetime):
    session = SessionLocal()
    seat = session.query(Seat).filter(Seat.id == seat_id).first()
    if not seat:
        session.close()
        return False
    
    duration = (end_time - start_time).total_seconds() / 60  # 转换为分钟
    is_valid = duration <= seat.max_duration
    session.close()
    return is_valid

def create_reservation(user_id: int, seat_id: int, start_time: datetime, end_time: datetime):
    if check_reservation_conflict(seat_id, start_time, end_time):
        raise ValueError("Seat already reserved for this time period")
    
    if not check_reservation_duration(seat_id, start_time, end_time):
        raise ValueError("Reservation duration exceeds maximum allowed time")
    
    session = SessionLocal()
    reservation = Reservation(
        user_id=user_id,
        seat_id=seat_id,
        start_time=start_time,
        end_time=end_time,
        status=ReservationStatus.ACTIVE
    )
    session.add(reservation)
    session.commit()
    session.refresh(reservation)
    
    # 创建提醒
    reminder_time = start_time - timedelta(minutes=30)  # 提前30分钟提醒
    create_reminder(reservation.id, user_id, reminder_time)
    
    session.close()
    return reservation

def get_reservations(library_id=None, user_id=None):
    session = SessionLocal()
    query = session.query(Reservation)
    if library_id:
        query = query.filter_by(library_id=library_id)
    if user_id:
        query = query.filter_by(user_id=user_id)
    reservations = query.all()
    session.close()
    return reservations

def cancel_reservation(reservation_id):
    session = SessionLocal()
    reservation = session.query(Reservation).filter_by(id=reservation_id).first()
    if reservation:
        reservation.status = ReservationStatus.CANCELLED
        session.commit()
    session.close()

def expire_reservations():
    session = SessionLocal()
    now = datetime.now()
    expired = session.query(Reservation).filter(
        and_(
            Reservation.status == ReservationStatus.ACTIVE,
            Reservation.end_time < now
        )
    ).all()
    for reservation in expired:
        reservation.status = ReservationStatus.EXPIRED
    session.commit()
    session.close()

def create_user(username, password, role="user"):
    session = SessionLocal()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    user = User(username=username, password=hashed_password, role=role)
    session.add(user)
    session.commit()
    session.close()

def verify_user(username, password):
    session = SessionLocal()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    user = session.query(User).filter_by(username=username, password=hashed_password).first()
    session.close()
    return user

def check_admin(user_id):
    session = SessionLocal()
    user = session.query(User).filter_by(id=user_id).first()
    session.close()
    return user and user.role == "admin"

def create_reminder(reservation_id, user_id, reminder_time):
    session = SessionLocal()
    reminder = Reminder(reservation_id=reservation_id, user_id=user_id, reminder_time=reminder_time)
    session.add(reminder)
    session.commit()
    session.close()

def send_reminders():
    session = SessionLocal()
    now = datetime.now()
    reminders = session.query(Reminder).filter(
        and_(
            Reminder.reminder_time <= now,
            Reminder.sent == False
        )
    ).all()
    for reminder in reminders:
        # 这里可以集成邮件/短信发送逻辑
        reminder.sent = True
    session.commit()
    session.close()

def update_reservation_stats():
    session = SessionLocal()
    now = datetime.now()
    for library in get_all_libraries():
        stats = ReservationStats(
            library_id=library.id,
            date=now,
            total_reservations=session.query(func.count(Reservation.id)).filter_by(library_id=library.id).scalar(),
            active_reservations=session.query(func.count(Reservation.id)).filter_by(library_id=library.id, status=ReservationStatus.ACTIVE).scalar(),
            cancelled_reservations=session.query(func.count(Reservation.id)).filter_by(library_id=library.id, status=ReservationStatus.CANCELLED).scalar(),
            expired_reservations=session.query(func.count(Reservation.id)).filter_by(library_id=library.id, status=ReservationStatus.EXPIRED).scalar()
        )
        session.add(stats)
    session.commit()
    session.close()

def create_area(library_id: int, name: str, description: str, floor: int):
    db = SessionLocal()
    area = Area(library_id=library_id, name=name, description=description, floor=floor)
    db.add(area)
    db.commit()
    db.refresh(area)
    db.close()
    return area

def create_seat(area_id: int, seat_number: str, seat_type: SeatType, has_power: bool = False, 
                has_computer: bool = False, max_duration: int = 120):
    db = SessionLocal()
    seat = Seat(
        area_id=area_id,
        seat_number=seat_number,
        seat_type=seat_type,
        has_power=has_power,
        has_computer=has_computer,
        max_duration=max_duration
    )
    db.add(seat)
    db.commit()
    db.refresh(seat)
    db.close()
    return seat

def get_area_seats(area_id: int):
    db = SessionLocal()
    seats = db.query(Seat).filter(Seat.area_id == area_id).all()
    db.close()
    return seats

def get_library_areas(library_id: int):
    db = SessionLocal()
    areas = db.query(Area).filter(Area.library_id == library_id).all()
    db.close()
    return areas

def get_seat_by_type(area_id: int, seat_type: SeatType):
    db = SessionLocal()
    seats = db.query(Seat).filter(
        and_(
            Seat.area_id == area_id,
            Seat.seat_type == seat_type
        )
    ).all()
    db.close()
    return seats

def get_available_seats_by_type(area_id: int, seat_type: SeatType):
    db = SessionLocal()
    seats = db.query(Seat).filter(
        and_(
            Seat.area_id == area_id,
            Seat.seat_type == seat_type,
            Seat.is_available == True
        )
    ).all()
    db.close()
    return seats

def get_seat_with_facilities(area_id: int, has_power: bool = None, has_computer: bool = None):
    db = SessionLocal()
    query = db.query(Seat).filter(Seat.area_id == area_id)
    
    if has_power is not None:
        query = query.filter(Seat.has_power == has_power)
    if has_computer is not None:
        query = query.filter(Seat.has_computer == has_computer)
    
    seats = query.all()
    db.close()
    return seats

def get_user_reservations(user_id: int, status: ReservationStatus = None):
    session = SessionLocal()
    query = session.query(Reservation).filter(Reservation.user_id == user_id)
    if status:
        query = query.filter(Reservation.status == status)
    reservations = query.all()
    session.close()
    return reservations

def get_seat_reservations(seat_id: int, status: ReservationStatus = None):
    session = SessionLocal()
    query = session.query(Reservation).filter(Reservation.seat_id == seat_id)
    if status:
        query = query.filter(Reservation.status == status)
    reservations = query.all()
    session.close()
    return reservations

def get_area_usage_stats(area_id: int, start_date: datetime, end_date: datetime):
    session = SessionLocal()
    stats = {
        "total_seats": session.query(func.count(Seat.id)).filter(Seat.area_id == area_id).scalar(),
        "reservations": {
            seat_type.value: session.query(func.count(Reservation.id)).join(Seat).filter(
                and_(
                    Seat.area_id == area_id,
                    Seat.seat_type == seat_type,
                    Reservation.start_time >= start_date,
                    Reservation.end_time <= end_date
                )
            ).scalar()
            for seat_type in SeatType
        }
    }
    session.close()
    return stats

def get_user_reservation_stats(user_id: int, start_date: datetime, end_date: datetime):
    """获取用户预约统计信息"""
    session = SessionLocal()
    try:
        stats = {
            "total": session.query(func.count(Reservation.id)).filter(
                and_(
                    Reservation.user_id == user_id,
                    Reservation.start_time >= start_date,
                    Reservation.end_time <= end_date
                )
            ).scalar(),
            "by_status": {
                status.value: session.query(func.count(Reservation.id)).filter(
                    and_(
                        Reservation.user_id == user_id,
                        Reservation.status == status,
                        Reservation.start_time >= start_date,
                        Reservation.end_time <= end_date
                    )
                ).scalar()
                for status in ReservationStatus
            }
        }
        return stats
    finally:
        session.close()

def log_system_action(level: LogLevel, module: str, action: str, details: str, user_id: int = None, ip_address: str = None):
    session = SessionLocal()
    log = SystemLog(
        level=level,
        module=module,
        action=action,
        details=details,
        user_id=user_id,
        ip_address=ip_address
    )
    session.add(log)
    session.commit()
    session.close()

def get_system_logs(start_time: datetime = None, end_time: datetime = None, level: LogLevel = None, module: str = None):
    session = SessionLocal()
    query = session.query(SystemLog)
    
    if start_time:
        query = query.filter(SystemLog.timestamp >= start_time)
    if end_time:
        query = query.filter(SystemLog.timestamp <= end_time)
    if level:
        query = query.filter(SystemLog.level == level)
    if module:
        query = query.filter(SystemLog.module == module)
    
    logs = query.order_by(SystemLog.timestamp.desc()).all()
    session.close()
    return logs

def import_seats_from_csv(area_id: int, csv_file_path: str, user_id: int = None):
    session = SessionLocal()
    imported_count = 0
    try:
        with open(csv_file_path, 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                seat = Seat(
                    area_id=area_id,
                    seat_number=row['seat_number'],
                    seat_type=SeatType[row['seat_type'].upper()],
                    has_power=row.get('has_power', '').lower() == 'true',
                    has_computer=row.get('has_computer', '').lower() == 'true',
                    max_duration=int(row.get('max_duration', 120))
                )
                session.add(seat)
                imported_count += 1
        session.commit()
        log_system_action(
            LogLevel.INFO,
            "seat_import",
            "import_seats",
            f"Imported {imported_count} seats to area {area_id}",
            user_id
        )
    except Exception as e:
        session.rollback()
        log_system_action(
            LogLevel.ERROR,
            "seat_import",
            "import_seats",
            f"Failed to import seats: {str(e)}",
            user_id
        )
        raise
    finally:
        session.close()
    return imported_count

def export_seats_to_csv(area_id: int, output_path: str):
    session = SessionLocal()
    seats = session.query(Seat).filter(Seat.area_id == area_id).all()
    
    with open(output_path, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(['seat_number', 'seat_type', 'has_power', 'has_computer', 'max_duration'])
        for seat in seats:
            writer.writerow([
                seat.seat_number,
                seat.seat_type.value,
                seat.has_power,
                seat.has_computer,
                seat.max_duration
            ])
    session.close()

def backup_database(backup_path: str):
    """创建数据库备份"""
    if os.path.exists("library.db"):
        shutil.copy2("library.db", backup_path)
        return True
    return False

def restore_database(backup_path: str):
    """从备份恢复数据库"""
    if os.path.exists(backup_path):
        shutil.copy2(backup_path, "library.db")
        return True
    return False

def batch_update_areas(library_id: int, area_updates: List[Dict[str, Any]], user_id: int = None):
    """批量更新区域信息"""
    session = SessionLocal()
    updated_count = 0
    try:
        for update in area_updates:
            area = session.query(Area).filter(Area.id == update['id']).first()
            if area and area.library_id == library_id:
                for key, value in update.items():
                    if key != 'id' and hasattr(area, key):
                        setattr(area, key, value)
                updated_count += 1
        session.commit()
        log_system_action(
            LogLevel.INFO,
            "area_management",
            "batch_update",
            f"Updated {updated_count} areas in library {library_id}",
            user_id
        )
    except Exception as e:
        session.rollback()
        log_system_action(
            LogLevel.ERROR,
            "area_management",
            "batch_update",
            f"Failed to update areas: {str(e)}",
            user_id
        )
        raise
    finally:
        session.close()
    return updated_count

def generate_token(length: int = 32) -> str:
    """生成随机token"""
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))

def create_user_with_verification(username: str, password: str, email: str, role: str = "user"):
    """创建用户并生成验证token"""
    session = SessionLocal()
    try:
        # 检查用户名和邮箱是否已存在
        if session.query(User).filter(User.username == username).first():
            raise ValueError("Username already exists")
        if session.query(User).filter(User.email == email).first():
            raise ValueError("Email already exists")
        
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        verification_token = generate_token()
        
        user = User(
            username=username,
            password=hashed_password,
            email=email,
            role=role,
            status=UserStatus.PENDING,
            verification_token=verification_token
        )
        session.add(user)
        session.commit()
        session.refresh(user)
        return user
    finally:
        session.close()

def verify_user_email(token: str) -> bool:
    """验证用户邮箱"""
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.verification_token == token).first()
        if user:
            user.status = UserStatus.ACTIVE
            user.verification_token = None
            session.commit()
            return True
        return False
    finally:
        session.close()

def create_password_reset_token(email: str) -> str:
    """创建密码重置token"""
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.email == email).first()
        if not user:
            return None
        
        reset_token = generate_token()
        user.reset_token = reset_token
        user.reset_token_expires = datetime.utcnow() + timedelta(hours=24)
        session.commit()
        return reset_token
    finally:
        session.close()

def reset_password(token: str, new_password: str) -> bool:
    """重置密码"""
    session = SessionLocal()
    try:
        user = session.query(User).filter(
            and_(
                User.reset_token == token,
                User.reset_token_expires > datetime.utcnow()
            )
        ).first()
        
        if user:
            user.password = hashlib.sha256(new_password.encode()).hexdigest()
            user.reset_token = None
            user.reset_token_expires = None
            session.commit()
            return True
        return False
    finally:
        session.close()

def add_to_blacklist(user_id: int, reason: str, end_date: datetime = None, admin_id: int = None):
    """将用户加入黑名单"""
    session = SessionLocal()
    try:
        # 检查用户是否已在黑名单中
        existing = session.query(Blacklist).filter(
            and_(
                Blacklist.user_id == user_id,
                or_(
                    Blacklist.end_date == None,
                    Blacklist.end_date > datetime.utcnow()
                )
            )
        ).first()
        
        if existing:
            raise ValueError("User is already blacklisted")
        
        blacklist = Blacklist(
            user_id=user_id,
            reason=reason,
            end_date=end_date,
            created_by=admin_id
        )
        session.add(blacklist)
        
        # 更新用户状态
        user = session.query(User).filter(User.id == user_id).first()
        if user:
            user.status = UserStatus.BLOCKED
        
        session.commit()
        return blacklist
    finally:
        session.close()

def remove_from_blacklist(user_id: int, admin_id: int = None):
    """将用户从黑名单中移除"""
    session = SessionLocal()
    try:
        blacklist = session.query(Blacklist).filter(
            and_(
                Blacklist.user_id == user_id,
                or_(
                    Blacklist.end_date == None,
                    Blacklist.end_date > datetime.utcnow()
                )
            )
        ).first()
        
        if blacklist:
            blacklist.end_date = datetime.utcnow()
            
            # 更新用户状态
            user = session.query(User).filter(User.id == user_id).first()
            if user:
                user.status = UserStatus.ACTIVE
            
            session.commit()
            return True
        return False
    finally:
        session.close()

def update_user_points(user_id: int, points: int, reason: str):
    """更新用户积分"""
    session = SessionLocal()
    try:
        user = session.query(User).filter(User.id == user_id).first()
        if not user:
            return False
        
        # 记录积分历史
        history = PointsHistory(
            user_id=user_id,
            points=points,
            reason=reason
        )
        session.add(history)
        
        # 更新用户积分
        user.points += points
        session.commit()
        return True
    finally:
        session.close()

def get_user_points_history(user_id: int, start_date: datetime = None, end_date: datetime = None):
    """获取用户积分历史"""
    session = SessionLocal()
    try:
        query = session.query(PointsHistory).filter(PointsHistory.user_id == user_id)
        
        if start_date:
            query = query.filter(PointsHistory.created_at >= start_date)
        if end_date:
            query = query.filter(PointsHistory.created_at <= end_date)
        
        return query.order_by(PointsHistory.created_at.desc()).all()
    finally:
        session.close()

def get_blacklist_status(user_id: int):
    """获取用户黑名单状态"""
    session = SessionLocal()
    try:
        blacklist = session.query(Blacklist).filter(
            and_(
                Blacklist.user_id == user_id,
                or_(
                    Blacklist.end_date == None,
                    Blacklist.end_date > datetime.utcnow()
                )
            )
        ).first()
        return blacklist
    finally:
        session.close()

def create_report(
    name: str,
    type: ReportType,
    format: ReportFormat,
    parameters: Dict,
    created_by: int
) -> Report:
    """创建报告任务"""
    session = SessionLocal()
    try:
        report = Report(
            name=name,
            type=type,
            format=format,
            parameters=parameters,
            created_by=created_by
        )
        session.add(report)
        session.commit()
        session.refresh(report)
        return report
    finally:
        session.close()

def update_report_status(
    report_id: int,
    status: str,
    file_path: str = None,
    error_message: str = None
) -> bool:
    """更新报告状态"""
    session = SessionLocal()
    try:
        report = session.query(Report).filter(Report.id == report_id).first()
        if report:
            report.status = status
            if file_path:
                report.file_path = file_path
            if error_message:
                report.error_message = error_message
            session.commit()
            return True
        return False
    finally:
        session.close()

def get_report(report_id: int) -> Optional[Report]:
    """获取报告信息"""
    session = SessionLocal()
    try:
        return session.query(Report).filter(Report.id == report_id).first()
    finally:
        session.close()

def get_user_reports(
    user_id: int,
    start_date: datetime = None,
    end_date: datetime = None
) -> List[Report]:
    """获取用户的报告列表"""
    session = SessionLocal()
    try:
        query = session.query(Report).filter(Report.created_by == user_id)
        if start_date:
            query = query.filter(Report.created_at >= start_date)
        if end_date:
            query = query.filter(Report.created_at <= end_date)
        return query.order_by(Report.created_at.desc()).all()
    finally:
        session.close()

def generate_excel_report(report: Report) -> str:
    """生成Excel报告"""
    try:
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "统计报告"
        
        # 设置样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 根据报告类型生成不同的内容
        if report.type == ReportType.DAILY:
            # 生成每日统计报告
            data = get_daily_stats(report.parameters.get("date"))
            headers = ["指标", "数值"]
            ws.append(headers)
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
            
            for key, value in data.items():
                ws.append([key, value])
        
        elif report.type == ReportType.WEEKLY:
            # 生成每周统计报告
            data = get_weekly_stats(report.parameters.get("week_start"))
            headers = ["日期", "预约总数", "活跃预约", "取消预约", "过期预约"]
            ws.append(headers)
            
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
            
            for row in data:
                ws.append(row)
        
        # 保存文件
        file_path = f"reports/{report.id}_{report.name}.xlsx"
        os.makedirs("reports", exist_ok=True)
        wb.save(file_path)
        return file_path
    
    except Exception as e:
        raise Exception(f"生成Excel报告失败: {str(e)}")

def generate_chart_data(report_type: ReportType, parameters: Dict) -> Dict:
    """生成图表数据"""
    try:
        if report_type == ReportType.DAILY:
            # 生成每日统计图表数据
            data = get_daily_stats(parameters.get("date"))
            return {
                "type": "bar",
                "data": {
                    "labels": list(data.keys()),
                    "datasets": [{
                        "label": "数值",
                        "data": list(data.values())
                    }]
                }
            }
        
        elif report_type == ReportType.WEEKLY:
            # 生成每周统计图表数据
            data = get_weekly_stats(parameters.get("week_start"))
            return {
                "type": "line",
                "data": {
                    "labels": [row[0] for row in data],
                    "datasets": [
                        {
                            "label": "预约总数",
                            "data": [row[1] for row in data]
                        },
                        {
                            "label": "活跃预约",
                            "data": [row[2] for row in data]
                        },
                        {
                            "label": "取消预约",
                            "data": [row[3] for row in data]
                        },
                        {
                            "label": "过期预约",
                            "data": [row[4] for row in data]
                        }
                    ]
                }
            }
    
    except Exception as e:
        raise Exception(f"生成图表数据失败: {str(e)}")

def get_system_config(key: str) -> Optional[SystemConfig]:
    """获取系统配置"""
    session = SessionLocal()
    try:
        return session.query(SystemConfig).filter(SystemConfig.key == key).first()
    finally:
        session.close()

def update_system_config(
    key: str,
    value: Any,
    description: str,
    updated_by: int
) -> SystemConfig:
    """更新系统配置"""
    session = SessionLocal()
    try:
        config = session.query(SystemConfig).filter(SystemConfig.key == key).first()
        if config:
            config.value = value
            config.description = description
            config.updated_by = updated_by
        else:
            config = SystemConfig(
                key=key,
                value=value,
                description=description,
                updated_by=updated_by
            )
            session.add(config)
        session.commit()
        session.refresh(config)
        return config
    finally:
        session.close()

def get_library_schedule(
    library_id: int,
    date: datetime = None
) -> List[LibrarySchedule]:
    """获取图书馆开放时间"""
    session = SessionLocal()
    try:
        query = session.query(LibrarySchedule).filter(LibrarySchedule.library_id == library_id)
        if date:
            query = query.filter(
                or_(
                    LibrarySchedule.special_date == date.date(),
                    and_(
                        LibrarySchedule.special_date == None,
                        LibrarySchedule.day_of_week == date.weekday()
                    )
                )
            )
        return query.all()
    finally:
        session.close()

def update_library_schedule(
    library_id: int,
    schedules: List[Dict],
    created_by: int
) -> List[LibrarySchedule]:
    """更新图书馆开放时间"""
    session = SessionLocal()
    try:
        # 删除现有的时间表
        session.query(LibrarySchedule).filter(
            LibrarySchedule.library_id == library_id
        ).delete()
        
        # 添加新的时间表
        new_schedules = []
        for schedule in schedules:
            new_schedule = LibrarySchedule(
                library_id=library_id,
                day_of_week=schedule["day_of_week"],
                open_time=schedule["open_time"],
                close_time=schedule["close_time"],
                is_closed=schedule.get("is_closed", False),
                special_date=schedule.get("special_date"),
                created_by=created_by
            )
            session.add(new_schedule)
            new_schedules.append(new_schedule)
        
        session.commit()
        for schedule in new_schedules:
            session.refresh(schedule)
        return new_schedules
    finally:
        session.close()

def get_reservation_rules(library_id: int) -> Optional[ReservationRule]:
    """获取预约规则"""
    session = SessionLocal()
    try:
        return session.query(ReservationRule).filter(
            ReservationRule.library_id == library_id
        ).first()
    finally:
        session.close()

def update_reservation_rules(
    library_id: int,
    rules: Dict,
    created_by: int
) -> ReservationRule:
    """更新预约规则"""
    session = SessionLocal()
    try:
        rule = session.query(ReservationRule).filter(
            ReservationRule.library_id == library_id
        ).first()
        
        if rule:
            for key, value in rules.items():
                setattr(rule, key, value)
            rule.updated_at = datetime.utcnow()
        else:
            rule = ReservationRule(
                library_id=library_id,
                created_by=created_by,
                **rules
            )
            session.add(rule)
        
        session.commit()
        session.refresh(rule)
        return rule
    finally:
        session.close()

def get_daily_stats(date: datetime) -> Dict:
    """获取指定日期的统计数据"""
    session = SessionLocal()
    try:
        # 获取指定日期的统计数据，如果不存在则返回默认值
        stats = session.query(ReservationStats).filter(func.date(ReservationStats.date) == date.date()).first()
        
        if stats:
            return {
                "总预约数": stats.total_reservations,
                "活跃预约数": stats.active_reservations,
                "已取消预约数": stats.cancelled_reservations,
                "已过期预约数": stats.expired_reservations
            }
        else:
            # 如果指定日期没有统计数据，返回0
            return {
                "总预约数": 0,
                "活跃预约数": 0,
                "已取消预约数": 0,
                "已过期预约数": 0
            }
    finally:
        session.close()

def get_weekly_stats(week_start: datetime) -> List[List[Any]]:
    """获取指定周的统计数据"""
    session = SessionLocal()
    try:
        week_end = week_start + timedelta(days=6)
        
        # 查询指定周内的每日统计数据
        stats = session.query(ReservationStats).filter(
            and_(
                func.date(ReservationStats.date) >= week_start.date(),
                func.date(ReservationStats.date) <= week_end.date()
            )
        ).order_by(ReservationStats.date).all()
        
        # 格式化数据为列表的列表
        weekly_data = []
        for day_stats in stats:
            weekly_data.append([
                day_stats.date.strftime('%Y-%m-%d'),
                day_stats.total_reservations,
                day_stats.active_reservations,
                day_stats.cancelled_reservations,
                day_stats.expired_reservations
            ])
        
        return weekly_data
    finally:
        session.close()

def get_notification_template(template_type: str) -> Optional[NotificationTemplate]:
    """获取通知模板"""
    session = SessionLocal()
    try:
        return session.query(NotificationTemplate).filter(NotificationTemplate.type == template_type).first()
    finally:
        session.close()

def get_pending_notifications():
    session = SessionLocal()
    notifications = session.query(Notification).filter_by(status=NotificationStatus.PENDING).all()
    session.close()
    return notifications

def process_notifications():
    """处理待发送的通知"""
    session = SessionLocal()
    try:
        pending_notifications = get_pending_notifications()
        for notification in pending_notifications:
            try:
                # 获取通知模板
                template = get_notification_template(notification.type.value)
                if not template:
                    notification.status = NotificationStatus.FAILED
                    notification.error_message = "Template not found"
                    continue

                # 准备通知内容
                # 这里需要根据实际情况替换模板中的变量
                # 例如，从 notification.metadata 中获取用户信息
                title = template.title_template.format(**notification.metadata or {})
                content = template.content_template.format(**notification.metadata or {})

                # 发送通知
                if notification.type == NotificationType.EMAIL:
                    # 发送邮件
                    # 这里需要实现发送邮件的逻辑
                    # 例如，使用 smtplib 发送邮件
                    pass
                elif notification.type == NotificationType.SMS:
                    # 发送短信
                    # 这里需要实现发送短信的逻辑
                    # 例如，调用短信服务商的 API
                    pass
                else:
                    notification.status = NotificationStatus.FAILED
                    notification.error_message = "Unsupported notification type"
                    continue

                # 更新通知状态
                notification.status = NotificationStatus.SENT
            except Exception as e:
                notification.status = NotificationStatus.FAILED
                notification.error_message = str(e)

        session.commit()
    finally:
        session.close() 